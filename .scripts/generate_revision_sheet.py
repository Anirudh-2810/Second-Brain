#!/usr/bin/env python3
"""Generate Interview Revision Sheet - Basic Python + LLM + 2 Repos (English only)"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = r"C:\Users\Vijaykumar\Second-Brain\Second-Brain\raw-sources\Interview_Revision_Sheet_Basic.xlsx"

# Styles
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FILL2 = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
HEADER_FILL3 = PatternFill(start_color="548235", end_color="548235", fill_type="solid")
HEADER_FILL4 = PatternFill(start_color="7F6000", end_color="7F6000", fill_type="solid")
SUB_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
ALT_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
WHITE_FILL = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin", color="B4C6E7"),
    right=Side(style="thin", color="B4C6E7"),
    top=Side(style="thin", color="B4C6E7"),
    bottom=Side(style="thin", color="B4C6E7"),
)
WRAP = Alignment(vertical="center", wrap_text=True, horizontal="left")
CENTER_WRAP = Alignment(vertical="center", wrap_text=True, horizontal="center")
HEADER_ALIGN = Alignment(vertical="center", wrap_text=True, horizontal="center")

COLS = [
    ("#", 6),
    ("Question", 42),
    ("30-sec Answer (Memorize)", 58),
    ("Example / Code", 38),
    ("Priority", 12),
    ("Status", 12),
]

def style_sheet(ws, header_fill, title, rows):
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_setup.orientation = "landscape"
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_options.horizontalCentered = True
    ws.oddHeader.center.text = title
    ws.oddHeader.center.font = "Calibri,Bold"
    ws.oddHeader.center.size = 10
    # column widths
    for idx, (h, w) in enumerate(COLS, 1):
        ws.column_dimensions[get_column_letter(idx)].width = w
    # header row
    for c, (h, w) in enumerate(COLS, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = header_fill
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER
    ws.row_dimensions[1].height = 22
    # freeze + filter
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:F{len(rows)+1}"
    # rows
    for r, row in enumerate(rows, 2):
        ws.row_dimensions[r].height = 45
        fill = ALT_FILL if r % 2 == 0 else WHITE_FILL
        for c, val in enumerate(row, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.font = Font(name="Calibri", size=10, color="1F1F1F")
            cell.fill = fill
            cell.border = THIN_BORDER
            if c == 1 or c == 5 or c == 6:
                cell.alignment = CENTER_WRAP
            else:
                cell.alignment = WRAP
            if c == 5:
                if val == "Must":
                    cell.fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
                    cell.font = Font(name="Calibri", bold=True, color="C00000", size=10)
                else:
                    cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
                    cell.font = Font(name="Calibri", color="375623", size=10)
    # status dropdown hint
    ws.sheet_properties.pageSetUpPr.fitToPage = True

# === DATA ===
python_rows = [
    [1, "What is Python? Key features?", "Interpreted, high-level, dynamically typed, indentation-based. Portable, large stdlib, garbage collected.", "python --version; print('hi')", "Must", "Todo"],
    [2, "Data types in Python?", "int, float, str, bool, list, tuple, dict, set, None. Use type(x) to check.", "type([1,2]) -> list", "Must", "Todo"],
    [3, "List vs Tuple vs Set vs Dict?", "List [] ordered+mutable+dup, Tuple () ordered+immutable, Set {} unordered+unique, Dict {k:v} key-value.", "a=[1,1]; set(a)->{1}", "Must", "Todo"],
    [4, "Mutable vs Immutable?", "Mutable can change after creation (list,dict,set). Immutable cannot (str,tuple,int).", "s='hi'; s[0]='H' -> error", "Must", "Todo"],
    [5, "== vs is?", "== checks value, is checks same object. Use is only for None.", "a==b True, a is b maybe False; x is None", "Must", "Todo"],
    [6, "Slicing?", "seq[start:stop:step]. s[::-1] reverses.", "s='abcd'; s[1:3]='bc'; s[::-1]='dcba'", "Must", "Todo"],
    [7, "List comprehension vs loop?", "Short way to build list: [x*2 for x in a if x>0]. Generator (x*2 for x in a) is lazy, low memory.", "[x*x for x in range(5)] -> [0,1,4,9,16]", "Should", "Todo"],
    [8, "dict.get vs d[key]?", "d[key] crashes if missing. d.get(k, default) safe.", "d.get('x',0) returns 0 if missing", "Must", "Todo"],
    [9, "Function + *args, **kwargs?", "def f(a,b=0):. *args tuple positional, **kwargs dict keyword.", "def f(*a,**k): print(a,k); f(1,2,x=3)", "Must", "Todo"],
    [10, "Mutable default arg bug?", "def f(a,l=[]) shares same list. Fix def f(a,l=None): if l is None: l=[]", "f(1)->[1], f(2)->[1,2] bug", "Must", "Todo"],
    [11, "Scope LEGB?", "Local > Enclosing > Global > Builtin. Use global/nonlocal to write outer.", "x=1; def f(): x=2 (local)", "Should", "Todo"],
    [12, "What is self and __init__?", "self = this object. __init__ runs on obj=Cls(). Stores data via self.name=name.", "class Dog: def __init__(self,n): self.n=n", "Must", "Todo"],
    [13, "Generator vs Iterator?", "Iterator has __next__. Generator uses yield, lazy, yields one at a time, saves memory.", "def gen(): yield 1; yield 2", "Should", "Todo"],
    [14, "with open? try/except?", "with auto-closes even on error. try/except catches errors, finally always runs.", "with open('a.txt') as f: data=f.read()", "Must", "Todo"],
    [15, "Decorator (basic)?", "Function wrapping function. @timer is f=timer(f). Use functools.wraps.", "@decorator\\ndef f(): pass", "Should", "Todo"],
    [16, "Shallow vs Deep copy? Copy vs alias?", "a=b alias same object. b=a[:] / copy() shallow top only. deepcopy recurses.", "import copy; b=copy.deepcopy(a)", "Should", "Todo"],
    [17, "GIL? Thread vs Process?", "GIL = one Python thread at a time. CPU-bound -> multiprocessing, IO-bound -> threading.", "CPU->Process, IO->Thread", "Should", "Todo"],
    [18, "venv + pip?", "venv per project isolated. pip install -r requirements.txt", "python -m venv .venv; pip freeze", "Should", "Todo"],
    [19, "Where are variables stored? (private heap)", "Variable is a name bound to an object. Objects live in private heap (Python memory manager). Name->object is a reference; id() shows address.", "a=5; b=a; id(a)==id(b) same object", "Must", "Todo"],
    [20, "Stack vs Heap in Python?", "Stack = call frames (refs). Heap = actual objects (lists, ints). Private heap + refcount + cyclic GC.", "def f(): x=[1,2] # [1,2] on heap", "Must", "Todo"],
    [21, "What is id() and refcount?", "Every object has id (addr in CPython), type, value, refcount. getrefcount() shows count; 0 -> freed.", "import sys; sys.getrefcount(a)", "Must", "Todo"],
    [22, "Why a is b True for small ints?", "CPython caches small ints -5..256 and interns some strings for speed; many names share one heap object. Don't use is for values.", "a=256; b=256 -> is True; 257 maybe False", "Must", "Todo"],
    [23, "How is Python dynamically typed?", "Type lives on object, not name. Name can rebond to any type.", "a=5; a='hi'; type(a) changes", "Must", "Todo"],
    [24, "What happens a=[1,2]; b=a; b.append(3)?", "a and b are two names to same heap list (alias). Mutate via one seen via other. Copy with b=a.copy() / a[:].", "a=[1,2]; b=a; b.append(3) -> a [1,2,3]", "Must", "Todo"],
    [25, "How does Python free memory? GC?", "Mostly refcount 0 -> freed now. Cycles a->b->a need cyclic GC (gc module). del a just drops one name.", "del a; import gc; gc.collect()", "Must", "Todo"],
    [26, "What are namespaces? Where?", "Namespace = name->object dict. Locals in frame (stack), globals in module dict, builtins in builtins dict. LEGB.", "locals(); globals()", "Should", "Todo"],
    [27, "Why everything is an object?", "Even int/func/class are heap objects with type/id/refcount. Allows a.__class__, func.__dict__.", "(5).__class__; f.x=1", "Should", "Todo"],
]

ml_rows = [
    [1, "What is ML? Types?", "ML learns pattern from data to predict. Types: Supervised (labeled), Unsupervised (no label), RL (agent+reward).", "spam yes/no = supervised", "Must", "Todo"],
    [2, "Classification vs Regression?", "Classification predicts category, Regression predicts number.", "cat/dog vs house price", "Must", "Todo"],
    [3, "Features vs Label?", "Features = inputs X, Label = output y to predict.", "X=[age,salary], y=buy?", "Must", "Todo"],
    [4, "Train / Test split? Why?", "Train 80% learn, Test 20% check unseen. Never test on train.", "train_test_split(test_size=0.2)", "Must", "Todo"],
    [5, "Overfit vs Underfit?", "Overfit memorizes train fails test. Underfit too simple fails both. Fix overfit: more data/simpler model.", "train 99/test 70 = overfit", "Must", "Todo"],
    [6, "Accuracy vs Precision/Recall?", "Accuracy=correct/total fails on imbalance. Precision=alarm trust, Recall=coverage. Cancer -> need Recall.", "95 fraud/5 real -> 95% acc useless", "Must", "Todo"],
    [7, "Confusion Matrix?", "Table TP,TN,FP,FN. All metrics from it.", "TP=hit, FP=false alarm", "Must", "Todo"],
    [8, "Cross-validation?", "Split train into k folds, rotate validation. More honest than single split.", "k=5 -> train 5 times", "Should", "Todo"],
    [9, "Scaling? When needed?", "Make features same range 0-1. Needed for KNN/gradient, not trees.", "StandardScaler", "Should", "Todo"],
    [10, "Basic algorithms (name + 1 line)?", "Linear/Logistic, Decision Tree (flowchart), Random Forest (many trees vote), KNN (neighbor vote), K-Means (clusters).", "Tree splits, Forest votes", "Must", "Todo"],
    [11, "Bias-Variance tradeoff?", "Error = bias2+variance+noise. Simple=high bias, Complex=high variance.", "Diagnose via learning curves", "Should", "Todo"],
    [12, "When does ZeroR/accuracy mislead?", "On imbalance. Use PR-AUC/F1 not accuracy.", "Fraud example above", "Should", "Todo"],
]

llm_rows = [
    [1, "What is LLM? Examples?", "Large Language Model trained to predict next token/word. Ex: ChatGPT, Gemini, Claude. Built on Transformer.", "ChatGPT = decoder-only", "Must", "Todo"],
    [2, "Transformer core idea?", "Every word looks at every other word via Attention, in parallel. No step-by-step reading.", "Attention = which words matter", "Must", "Todo"],
    [3, "Self-Attention formula?", "Attention=softmax(QK^T/sqrt(dk)) V where Q=XWq, K=XWk, V=XWv. Scale stops saturation.", "Write it on board", "Must", "Todo"],
    [4, "Multi-head attention?", "h heads parallel each learns different relation, concat + project.", "h=8 heads", "Should", "Todo"],
    [5, "Why positional encoding?", "Attention has no order. Add sin/cos or learned position vector to give order.", "PE = sin(pos/10000...)", "Should", "Todo"],
    [6, "Token? Embedding?", "Token = piece of word (BPE). Embedding = word->vector numbers; similar words near.", "'ChatGPT'->'Chat'+'GPT'", "Must", "Todo"],
    [7, "How LLM trained? (3 steps)", "1) Pre-train next-token on web 2) SFT instruction pairs 3) RLHF human ratings.", "Pretrain->SFT->RLHF", "Must", "Todo"],
    [8, "Temperature / Top-p?", "Temperature 0=deterministic, >1 creative. Top-p samples smallest set summing to p.", "temp 0.7 balanced", "Must", "Todo"],
    [9, "Context window? O(n2)?", "Max tokens model sees. Attention cost O(T^2) double length = 4x compute.", "4K vs 128K window", "Should", "Todo"],
    [10, "Hallucination? RAG?", "Hallucination = fluent false. RAG = search docs -> feed to LLM to ground answer, reduces hallucination.", "Docs + LLM = RAG", "Must", "Todo"],
    [11, "Fine-tuning vs Prompt engineering?", "Prompt engineering = better instructions. Fine-tuning/LoRA = train small adapter on your data.", "LoRA trains <1% params", "Should", "Todo"],
    [12, "Encoder vs Decoder?", "Encoder bidirectional (BERT). Decoder causal masked sees <=t (GPT). LLMs = decoder-only.", "GPT = decoder-only", "Should", "Todo"],
]

pomodoro_rows = [
    [1, "What is Quote Pomodoro? Stack?", "Tkinter desktop Pomodoro timer, 25/5,50/10,15/3 presets, quotes, beeps, plyer notifications. Python tkinter/threading/winsound.", "flightproductivity.py 198 lines", "Must", "Todo"],
    [2, "Architecture: threading model?", "Main thread = UI (mainloop). Daemon thread = timer loop sleep(1)->remaining--. Bridge via root.after(0,tick_ui) thread-safe.", "Thread(target=loop, daemon=True)", "Must", "Todo"],
    [3, "Why root.after not direct UI update?", "Tkinter NOT thread-safe. Direct update from timer thread crashes. root.after schedules tick_ui on main thread.", "self.root.after(0, self.tick_ui)", "Must", "Todo"],
    [4, "Pause/Resume how?", "Flag is_paused. Loop: if is_paused: continue (skip decrement). Toggle button text Pause<->Resume.", "is_paused not is_paused", "Must", "Todo"],
    [5, "Presets + progress bar?", "PRESETS dict (25*60 etc). Progress = ttk.Progressbar maximum=total, value=total-remaining each sec.", "PRESETS[\"25/5\"]=(1500,300)", "Should", "Todo"],
    [6, "Sounds cross-platform?", "Windows winsound.Beep(800,160) start, 600+450 end. Else print('\\a') terminal bell. Wrapped in try.", "HAS_WINSOUND check", "Should", "Todo"],
    [7, "Quotes rotation?", "10 quotes list, index increments every 300s via time.time() check in loop, modulo wrap.", "QUOTES[(i+1)%len]", "Should", "Todo"],
    [8, "Hardest part / gotcha to say?", "Thread-safety + pause state + editable entry parsing mm:ss with fallback. Fix: root.after + flags + try parse.", "Talk in STAR format", "Must", "Todo"],
    [9, "What would you improve?", "CSV logging sessions, task name field, auto-break toggle, stats dashboard, system tray.", "Known TODOs in vault", "Should", "Todo"],
]

handsens_rows = [
    [1, "What is handsens101? Stack?", "Webcam hand-gesture mouse. Python, OpenCV, MediaPipe HandLandmarker, pyautogui. Repo Anirudh-2810/handsens101.", "src/main.py JarvisUltimaPro", "Must", "Todo"],
    [2, "Pipeline step-by-step?", "OpenCV cap 0 -> MediaPipe (1 hand, conf 0.85) -> 21 landmarks -> map to screen px -> exponential smoothing 5.0 -> pyautogui move/click/scroll.", "detect->filter->map->actuate", "Must", "Todo"],
    [3, "Gestures mapped?", "Pinch index+thumb = click, Index+middle together = scroll, Normal move = cursor.", "Pinch->click, 2-finger->scroll", "Must", "Todo"],
    [4, "Why confidence 0.85?", "Raised from default to cut false positives/noise. Reduces jitter detections.", "HandLandmarker conf 0.85", "Should", "Todo"],
    [5, "Why smoothing 5.0?", "Raw landmarks jitter. Exponential smoothing cursor=prev*alpha+new*(1-alpha) stabilizes before actuation.", "Same as robotics filtering", "Must", "Todo"],
    [6, "Why pyautogui PAUSE=0 failsafe off?", "Low latency. PAUSE adds delay per call. Failsafe off avoids corner-stop. Tradeoff: less safety.", "pyautogui.PAUSE=0", "Should", "Todo"],
    [7, "Hardest part / failure?", "Jitter + lighting noise. Fixed via confidence + smoothing. Also latency vs accuracy tradeoff.", "Demo in good light", "Must", "Todo"],
    [8, "How to extend?", "ROS2 teleop, add gesture classification, 2-hand support, depth control, calibration for screen mapping.", "Robotics stack mini", "Should", "Todo"],
]

# Build workbook
wb = openpyxl.Workbook()

# Remove default sheet
wb.remove(wb.active)

# Helper to create sheet
def add_sheet(name, header_fill, title, rows):
    ws = wb.create_sheet(title=name)
    style_sheet(ws, header_fill, title, rows)
    return ws

# Order: How to use first
# Create sheets in desired order
ws_intro = add_sheet("00_How_To_Use", HEADER_FILL, "How To Use - Read First", [
    [1, "How to use this sheet?", "Filter Priority=Must for tonight. Set Status Done as you practice aloud. 1 row = 30-sec answer. Don't read verbatim — use skeleton.", "Interview tactic: Definition->Why->When breaks", "Must", "Todo"],
    [2, "Scoring tip (ML Playbook)", "Interviewers probe till boundary. Say boundary fast: 'Derivation escapes me, but result is X' -> narrate reasoning. Silence = fail.", "wiki/01-Areas/AI-Data/data-science/ml-interview-playbook.md", "Must", "Todo"],
])
# Adjust intro rows height a bit bigger
for r in range(2, ws_intro.max_row+1):
    ws_intro.row_dimensions[r].height = 55

add_sheet("01_Python_Basic", HEADER_FILL, "Python Basic - 27 Qs", python_rows)
add_sheet("02_ML_Basic", HEADER_FILL2, "ML Basic - 12 Qs", ml_rows)
add_sheet("03_LLM_Basic", HEADER_FILL3, "LLM Basic - 12 Qs", llm_rows)
add_sheet("04_Repo_Pomodoro", PatternFill(start_color="404040", end_color="404040", fill_type="solid"), "Repo: Quote Pomodoro - 9 Qs", pomodoro_rows)
add_sheet("05_Repo_Handsens101", PatternFill(start_color="7F6000", end_color="7F6000", fill_type="solid"), "Repo: handsens101 - 8 Qs", handsens_rows)
# New: Frontend basics
frontend_rows = [
    [1, "HTML vs CSS vs JS roles?", "HTML structure, CSS presentation, JS behavior. Together render page.", "<button> + button{color} + onclick", "Must", "Todo"],
    [2, "Semantic HTML? Why?", "Use header/nav/main/section not just divs - SEO + a11y + readability.", "<nav><ul> not <div><div>", "Should", "Todo"],
    [3, "Box model?", "content+padding+border+margin. Use box-sizing:border-box to fix width math.", "div{box-sizing:border-box}", "Must", "Todo"],
    [4, "Flex vs Grid?", "Flex 1D (row/col) for nav; Grid 2D for layout. Both for responsive.", "display:flex; justify-content:center", "Must", "Todo"],
    [5, "Specificity?", "Rank: inline > id > class > tag. !important breaks it. Avoid high specificity.", "#id .class tag", "Should", "Todo"],
    [6, "let vs const vs var?", "var function-scoped hoisted (avoid). let block mutable, const binding immutable (object still mutates).", "const a=[1]; a.push(2) ok", "Must", "Todo"],
    [7, "DOM + event handling?", "DOM is tree of nodes JS manipulates. Add listeners for interaction.", "querySelector('#btn').addEventListener('click',fn)", "Must", "Todo"],
    [8, "fetch + async/await?", "fetch returns Promise. Use await fetch(url) then .json(). Handle CORS/errors.", "await fetch('/api/quotes')", "Must", "Todo"],
    [9, "Responsive design?", "viewport meta + fluid units (rem/%) + media queries. Mobile-first.", "@media(max-width:600px){...}", "Should", "Todo"],
    [10, "Basic accessibility?", "alt on images, label for inputs, keyboard focus, contrast.", '<img alt="quote">', "Should", "Todo"],
]
add_sheet("06_Frontend_Basic", PatternFill(start_color="9E4D2E", end_color="9E4D2E", fill_type="solid"), "Frontend Basic - 10 Qs", frontend_rows)

# Set tab colors
wb["00_How_To_Use"].sheet_properties.tabColor = "1F4E78"
wb["01_Python_Basic"].sheet_properties.tabColor = "1F4E78"
wb["02_ML_Basic"].sheet_properties.tabColor = "2E75B6"
wb["03_LLM_Basic"].sheet_properties.tabColor = "548235"
wb["04_Repo_Pomodoro"].sheet_properties.tabColor = "404040"
wb["05_Repo_Handsens101"].sheet_properties.tabColor = "7F6000"
wb["06_Frontend_Basic"].sheet_properties.tabColor = "9E4D2E"

# Print titles
for ws in wb.worksheets:
    ws.print_title_rows = "1:1"

wb.save(OUT)
print(f"SAVED -> {OUT}")
# Verify
wb2 = openpyxl.load_workbook(OUT)
print("Sheets:", wb2.sheetnames)
for n in wb2.sheetnames:
    print(n, "rows", wb2[n].max_row, "cols", wb2[n].max_column)
